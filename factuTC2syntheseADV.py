# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os

racine=r"c:\users\sdecaluwe\desktop\factuTC"
path_lut = racine+"\LUT.xlsx"

wb = Workbook()

wb = load_workbook(filename = path_lut)

# aller dans mapping retail TB

ws = wb["matrice clients TEM"]

dico_client={}

for row in range(ws.max_row+1):
   code_client_SAP= str(ws.cell(row+1,3).value)
   client=str(ws.cell(row+1,1).value)
   dico_client[client]=code_client_SAP


# effacement de SAP/TetraConnect.xlsx
for i in os.listdir(racine):
   if i.endswith("TetraConnect.xlsx"):
      os.remove(racine+"/"+i)


# caler colonnes (synthèses ADV, regarder import lut pour client => code SAP
# convertir en .xslx ?
# effacer fichier précédent (regarder conversion SAP)
# interfacer avec conversion SAP

path_file="C:/Users/sdecaluwe/Desktop/factuTC/20220404_15h04_factuGlobale_Q1_2022.xlsx"
nom_fichier="C:/Users/sdecaluwe/Desktop/factuTC/syntheseADV-TetraConnect- Q1 2022.xlsx"

###
wb_input=Workbook()
wb_input = load_workbook(filename=path_file)
ws_input = wb_input["Clients sans tri TEM-TC"]
for i in range(1,10):
   row=[]
   for j in range(1,10):
      row.append(ws_input.cell(i,j).value)
      

wb_output = Workbook(write_only=True)
ws_output = wb_output.create_sheet('Synthèse ADV')


#for row in rows_sim_passport:
#    if len(row) > 6:
#        if row[3]!="Contract":
#            ws_output.append(row)
#            #print(row)
#wb_output_sim_passport.save("resultat_sim_passport.xlsx")
###




codeSAP=[0]*(ws.max_column+1)
row_title=["code SAP","client","identifiants","Activité","Opération","Logiciel","Article de prestation","Libellé","Quantité","Prix","Montant","Entre","et"]
ws_output.append(row_title)

for row in range(2, ws_input.max_row+1):
   for column in range(1, ws_input.max_column+1):
       cell = str(ws_input.cell(row,column).value)
       if column==1:
          client = cell
       else:
          quantity = cell
          libelle = ws_input.cell(1,column).value
          codeSAP=dico_client[client]
          if quantity != "None" and quantity !="0":
             row_to_add=[codeSAP, client, "", "service", "service", "",libelle,"", quantity, "", "","01/01/2022","31/03/2022"]
             print(row_to_add)
             ws_output.append(row_to_add)
wb_output.save(nom_fichier)







  
